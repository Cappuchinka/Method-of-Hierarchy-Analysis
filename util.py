import math
import csv
import os
from fractions import Fraction

import numpy as np
import pandas as pd
from IPython.display import display, Markdown, HTML

pd.set_option('display.precision', 3)
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Font, Alignment
from openpyxl.utils import get_column_letter


def parse_value(value):
    value = value.strip()
    if not value:
        return None
    if '/' in value:
        numerator, denominator = value.split('/')
        return Fraction(int(numerator), int(denominator))
    else:
        return int(value)


def format_fraction(x):
    return f"{x.real}" if isinstance(x, Fraction) else str(x)


def get_matrix_paired_comparisons(data, names):
    df = pd.DataFrame(data, index=names, columns=names)

    return df


def get_priority_vectors(data, names):
    matrix = np.array([[float(item) for item in row] for row in data])

    priority_vector = np.exp(np.mean(np.log(matrix), axis=1))
    total_sum = np.sum(priority_vector)
    priority_vector /= round(total_sum, 3)

    return pd.DataFrame(priority_vector, index=names, columns=['Вектор приоритетов'])


def get_lambda_max(matrix, vector):
    np_matrix = np.array(matrix)
    np_vector = np.array(vector)
    matrix_dot_vector = np.dot(np_matrix, np_vector)

    return (1 / len(np_matrix)) * np.sum(matrix_dot_vector / vector, axis=0)


def get_consistency_index(param_lambda_max, quantity):
    return math.copysign(((param_lambda_max - quantity) / (quantity - 1)), 1)


def get_random_consistency_index(n):
    random_consistency_indexes = np.array([0, 0, 0.58, 0.90, 1.12, 1.24, 1.32, 1.41, 1.45, 1.49])
    return random_consistency_indexes[n - 1]


def get_consistency_relation(param_consistency_index, n):
    return math.copysign(param_consistency_index / get_random_consistency_index(n), 1)


def count_global_priorities(p_v_criteria, p_vs_alternatives):
    quantity_alternatives = len(p_vs_alternatives[0])
    quantity_criteria = len(p_v_criteria)
    priorities = []

    for j in range(0, quantity_alternatives):
        total = 0
        for i in range(0, quantity_criteria):
            total += p_vs_alternatives[i][j] * p_v_criteria[i]
        priorities.append(total)

    return priorities


def save_matrix_paired_comparisons_to_xlsx(data, val_1, val_2, val_3, name):
    try:
        df = data.copy()

        times_new_roman = Font(name='Times New Roman', size=12)
        alignment = Alignment(wrap_text=True, vertical='center')

        def format_value(x):
            if hasattr(x, 'numerator'):
                if x.denominator == 1:
                    return str(x.numerator)
                return f"{x.numerator}/{x.denominator}"
            elif isinstance(x, (float, int)):
                return f"{float(x):.3f}"
            return str(x)

        def wrap_words(text):
            if not isinstance(text, str):
                return text

            return text.replace(' ', '\n')

        wb = Workbook()
        ws = wb.active

        ws.cell(row=1, column=1, value="").font = times_new_roman

        for col_idx, col_name in enumerate(df.columns, 2):
            wrapped_name = wrap_words(str(col_name))
            cell = ws.cell(row=1, column=col_idx, value=wrapped_name)
            cell.font = times_new_roman
            cell.alignment = alignment

        for row_idx, (index, row) in enumerate(df.iterrows(), 2):
            wrapped_index = wrap_words(str(index))
            cell = ws.cell(row=row_idx, column=1, value=wrapped_index)
            cell.font = times_new_roman
            cell.alignment = alignment

            for col_idx, value in enumerate(row, 2):
                wrapped_value = wrap_words(format_value(value))
                cell = ws.cell(row=row_idx, column=col_idx, value=wrapped_value)
                cell.font = times_new_roman
                cell.alignment = alignment

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter

            for cell in col:
                try:
                    if cell.value and '\n' in str(cell.value):
                        line_length = max(len(line) for line in str(cell.value).split('\n'))
                    else:
                        line_length = len(str(cell.value))

                    if line_length > max_length:
                        max_length = line_length
                except:
                    pass

            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width

        last_col = df.shape[1] + 1
        start_row = df.shape[0] + 2

        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        ws.merge_cells(start_row=start_row, start_column=1,
                       end_row=start_row + 2, end_column=last_col - 1)

        ws.cell(row=start_row, column=1, value="").font = times_new_roman

        cell = ws.cell(row=start_row, column=last_col, value=f"λ_max = {val_1:.3f}")
        cell.font = times_new_roman
        cell.alignment = alignment
        cell = ws.cell(row=start_row + 1, column=last_col, value=f"ИС = {val_2:.3f}")
        cell.font = times_new_roman
        cell.alignment = alignment
        cell = ws.cell(row=start_row + 2, column=last_col, value=f"ОС = {val_3:.3f}")
        cell.font = times_new_roman
        cell.alignment = alignment

        for row in ws.iter_rows(min_row=1, max_row=start_row + 2, max_col=last_col):
            for cell in row:
                cell.border = thin_border
                cell.font = times_new_roman
                cell.alignment = alignment

        os.makedirs("out", exist_ok=True)
        file_path = os.path.join("out", f"{name}.xlsx")
        wb.save(file_path)

        display(Markdown(f"Файл успешно сохранён: {file_path}"))
        return True

    except Exception as e:
        display(Markdown(f"Ошибка при сохранении файла: {e}"))
        return False


def save_global_priorities_to_excel(data, name):
    try:
        if not isinstance(data, pd.DataFrame):
            raise ValueError("Входные данные должны быть pandas DataFrame")

        if len(data) == 0:
            raise ValueError("DataFrame не содержит данных")

        def wrap_words(text):
            if not isinstance(text, str):
                return text

            return text.replace(' ', '\n')

        wb = Workbook()
        ws = wb.active
        ws.title = "Данные"

        font = Font(name='Times New Roman', size=12)
        border = Border(left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin'))
        alignment = Alignment(wrap_text=True, vertical='center')

        def format_value(value):
            try:
                if isinstance(value, (int, float)):
                    if isinstance(value, int) or value.is_integer():
                        return str(int(value))
                    return f"{float(value):.3f}"
                return str(value)
            except:
                return str(value)

        cell = ws.cell(row=1, column=1, value="")
        cell.font = font
        cell.border = border
        cell.alignment = alignment

        for col_num, column_name in enumerate(data.columns, 2):
            wrapped_name = wrap_words(str(column_name))
            cell = ws.cell(row=1, column=col_num, value=wrapped_name)
            cell.font = font
            cell.border = border
            cell.alignment = alignment

        for row_num, (index, row) in enumerate(data.iterrows(), 2):
            wrapped_index = wrap_words(str(index))
            cell = ws.cell(row=row_num, column=1, value=wrapped_index)
            cell.font = font
            cell.border = border
            cell.alignment = alignment

            for col_num, value in enumerate(row, 2):
                wrapped_value = wrap_words(format_value(value))
                cell = ws.cell(row=row_num, column=col_num, value=wrapped_value)
                cell.font = font
                cell.border = border
                cell.alignment = alignment

        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)

            for cell in col:
                try:
                    if cell.value and '\n' in str(cell.value):
                        line_length = max(len(line) for line in str(cell.value).split('\n'))
                    else:
                        line_length = len(str(cell.value))

                    if line_length > max_length:
                        max_length = line_length
                except:
                    pass

            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[col_letter].width = adjusted_width

        os.makedirs("out", exist_ok=True)
        file_path = os.path.join("out", f"{name}.xlsx")
        wb.save(file_path)

        display(Markdown(f"Файл успешно сохранён: {file_path}"))
        return True

    except Exception as e:
        display(Markdown(f"Ошибка при сохранении: {str(e)}"))
        return False